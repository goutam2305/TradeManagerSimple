import openpyxl

def search_broad(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                val = str(cell.value).upper() if cell.value else ""
                if "STOP" in val or "LOSS" in val or "WIN" in val or "CAPITAL" in val:
                    print(f"[{sheet_name}] {cell.coordinate}: {cell.value}")

if __name__ == "__main__":
    search_broad(r"d:\Test\TradeManagerSimple\Reference\money-management-sheet.xlsx")
