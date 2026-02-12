import openpyxl

def dump_main_page(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb["MAIN PAGE"]
    with open("main_page_dump.txt", "w", encoding="utf-8") as f:
        for r in range(1, 51):
            row_vals = []
            for c in range(1, 10):
                cell = sheet.cell(row=r, column=c)
                row_vals.append(f"{cell.coordinate}: {cell.value}")
            f.write(" | ".join(row_vals) + "\n")
    
    wb_vals = openpyxl.load_workbook(file_path, data_only=True)
    sheet_v = wb_vals["MAIN PAGE"]
    with open("main_page_values.txt", "w", encoding="utf-8") as f:
        for r in range(1, 51):
            row_vals = []
            for c in range(1, 10):
                cell = sheet_v.cell(row=r, column=c)
                row_vals.append(f"{cell.coordinate}: {cell.value}")
            f.write(" | ".join(row_vals) + "\n")

if __name__ == "__main__":
    dump_main_page(r"d:\Test\TradeManagerSimple\Reference\money-management-sheet.xlsx")
